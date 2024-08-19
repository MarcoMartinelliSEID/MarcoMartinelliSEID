#----------------------------------------------------------------------------------------------------------------------------------------------------------------------
#SEID Srl
#Progetto: Function Generator
#Versione: 1.0
#Data: 10/06/2020
#Aggiornamenti: xxx
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------

from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog
import time
from BitToPLC import Segnale
from BitToPLC import Gestore
import tkinter as tkinter
import os


filename = 'Empty'

def openFunctionGen():

    def runFunctionGen():
        
        #codice della classe TagNumber (spostabile su file a parte)
        class TagNumber:
            tag = None
            analog = None
            status = None
            func = None

            def __init__(self, tag, analog, status, func): #costruttore dell'oggetto
                self.tag = tag
                self.analog = analog
                self.status = status
                self.func = func
            
            def print(self):
                print(str(self.analog) + "     " + str(self.status) + "    " + self.tag)
                text.insert(INSERT, (str(self.analog) + "     " + str(self.status) + "    " + self.tag + '\n'))

            @staticmethod
            def addNetwork(dati): #Metodo statico per la creazione di NETWORK da inserire
                if dati.status != 'EMPTY' and dati.analog != 'EMPTY': #se il tag esiste sia ANALOG che STATUS
                    return ('NETWORK\nTITLE =\n\n' + 
                            '        L        "' + dati.tag + '".Val;\n' +
                            '        T        "' + func + '".READ.ANALOG.r' + dati.analog + ';\n\n' +
                            '        L        "' + dati.tag + '".StsDcs;\n' +
                            '        T        "' + func + '".READ.STATUS.r' + dati.status + ';\n\n')
                else:
                    if dati.analog != 'EMPTY': #se il tag esiste solo ANALOG
                        return ('NETWORK\nTITLE =\n\n' + 
                            '        L        "' + dati.tag + '".Val;\n' +
                            '        T        "' + func + '".READ.ANALOG.r' + dati.analog + ';\n\n')
                    else: #se il tag esiste solo STATUS
                        return ('NETWORK\nTITLE =\n\n' + 
                            '        L        "' + dati.tag + '".StsDcs;\n' +
                            '        T        "' + func + '".READ.STATUS.r' + dati.status + ';\n\n')
                return '\n------------------ SytaxError --------------------\n\n'
                    
        #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        print("Processing...")
        text.insert(INSERT, "Processing...")
        label_elabora.configure(text = "Processing...")

        #lettura file.xlsx
        global filename
        fileIN = load_workbook(filename)
        sheet = fileIN["Foglio1"]

        #creazione file.awl
        fileOUT = open("SiemensFC.awl", "w+") #C:\\ ... tra le virgolette, per accedere a file in altri folder

        #+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        #Intestazione file
        fileOUT.write('FUNCTION ' + str(sheet.cell(1,2).value) + ' : VOID\n')
        fileOUT.write('TITLE = \n')
        fileOUT.write('VERSION : 0.1\nBEGIN\n')
        print('\nANALOG    STATUS   TAG\n')
        text.insert(INSERT, '\n\nANALOG    STATUS   TAG\n\n')

        #+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        #MAIN PROGRAM
        array = []
        xstatus = 4
        ystatus = 1
        x = 4
        y = 1

        #ciclo di lettura dati ANALOG e verifica che siano anche STATUS
        while str.upper(sheet.cell(x, y).value) == 'ANALOG':
            tag = str(sheet.cell(x, y+2).value)
            analog = str(sheet.cell(x, y+1).value)
            func = str(sheet.cell(1, 2).value)
            status = 'EMPTY'
            xstatus = 4

            while str.upper(sheet.cell(xstatus, ystatus).value) != 'STATUS': #scorro fino alla sezione STATUS
                xstatus += 1

            while str.upper(sheet.cell(xstatus, ystatus).value) == 'STATUS':
                if str(sheet.cell(xstatus, ystatus+2).value) == tag:
                    status = str(sheet.cell(xstatus, ystatus+1).value)
                xstatus += 1
                    
            
            if (sheet.cell(x, y+2).value) is not None:
                dati = TagNumber(tag, analog, status, func) #creazione oggetto della classe TagNumber
                fileOUT.write(TagNumber.addNetwork(dati))
                array.append(dati) #aggiunta tag nell'array che serve successivamente nel ciclo degli STATUS per controllare che un certo tag STATUS sia già stato trovato tra gli ANALOG
                dati.print() #print serve solo per avere un feedback sul terminale
            x += 1


        xstatus = 4
        ystatus = 1

        #ciclo di lettura dati STATUS che non sono ANALOG
        while str.upper(sheet.cell(xstatus, ystatus).value) != 'STATUS':
            xstatus += 1

        while str.upper(sheet.cell(xstatus, ystatus).value) == 'STATUS':
            tag = str(sheet.cell(xstatus, ystatus+2).value)
            status = str(sheet.cell(xstatus, ystatus+1).value)
            func = str(sheet.cell(1, 2).value)
            analog = 'EMPTY'
                
            i = 0
            test = False
            l = len(array)
            #ciclo che verifica che il tag STATUS preso in considerazione non sia già stato trovato in ANALOG nel primo ciclo
            while i < l:
                if array[i].tag == tag:
                    test = True
                i += 1        
                
            if (sheet.cell(xstatus, ystatus+2).value) is not None and test == False:
                dati = TagNumber(tag, analog, status, func)
                fileOUT.write(TagNumber.addNetwork(dati))
                dati.print()
            xstatus += 1

        print('\nDone.')
        label_elabora.configure(text = "Done.")
        label_output.configure(text = "File Output Pronto")
        text.insert(INSERT, '\n\nDone.\n')
        text.see("end")



    def browseFiles():
        global filename
        filename = filedialog.askopenfilename(initialdir = "/", 
                                            title = "Seleziona il file excel", 
                                            filetypes = (("Excel Files", "*.xlsx*"), ("Excel Files", "*.xlsm*"), ("all files", "*.*")))

        label_file_explorer.configure(text = "File aperto: " + filename)
        label_elabora.configure(text = "Pronto")

    def openOutputFile():
        os.system("start " + 'SiemensFC.awl')



    #OPEN WINDOW
    FunctGen = Tk()
    FunctGen.iconbitmap('img/seid.ico')
    logo = PhotoImage(file="img/SEIDN3.jpg")
            
    #FINESTRA
    FunctGen.title('SEID File Explorer')
    FunctGen.geometry("700x700")
    FunctGen.config(background = "black")

    # TITOLO SCHERMATA
    label_file_explorer = Label(FunctGen, text = "Nessun File Aperto", width = 80, height = 4, fg = "white", background = "#404040")   
    label_elabora = Label(FunctGen, text = "Attesa", width = 80, height = 4, fg = "white", background = "#404040")
    label_output = Label(FunctGen, text = "", width = 80, height = 4, fg = "white", background = "#404040")
    label_exit = Label(FunctGen, text = "Exit", width = 80, height = 4, fg = "white", background = "#404040")
    labelLogo = Label(FunctGen, image=logo, bd = 0)

    #BUTTON
    button_esplora = Button(FunctGen, text = "Open File", command = browseFiles, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')  
    button_elabora = Button(FunctGen, text = "Elabora", command = runFunctionGen, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')
    button_output = Button(FunctGen, text = "Open Output File", command = openOutputFile, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')
    button_exit = Button(FunctGen, text = "Chiudi", command = exit, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')

    #Struttura
    label_file_explorer.grid(column = 2, row = 1)
    label_elabora.grid(column = 2, row = 2)
    label_output.grid(column = 2, row = 3)
    label_exit.grid(column = 2, row = 4)
    button_esplora.grid(column = 1, row = 1)
    button_elabora.grid(column = 1, row = 2)
    button_output.grid(column = 1, row = 3)
    button_exit.grid(column = 1, row = 4)
    labelLogo.grid(column = 1, row = 5)

    #scrollText
    frame = Frame(FunctGen, width=385, height=300, bd=0, relief=SUNKEN)
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)
    xscrollbar = Scrollbar(frame, orient=HORIZONTAL)
    #xscrollbar.grid(row=1, column=0, sticky=E+W)
    yscrollbar = Scrollbar(frame)
    #yscrollbar.grid(row=0, column=1, sticky=N+S)
    text = Text(frame, width=70, height=24, wrap=NONE, bd=0, xscrollcommand=xscrollbar.set, yscrollcommand=yscrollbar.set, fg = "#4bc973", background = "black")
    text.grid(row=0, column=0, sticky=N+S+E+W)
    xscrollbar.config(command=text.xview)
    yscrollbar.config(command=text.yview)
    frame.grid(row=5, column=2)

    #loop window
    FunctGen.mainloop()

