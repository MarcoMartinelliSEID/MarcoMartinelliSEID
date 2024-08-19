import time
from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog, font
import tkinter as tk
import os.path
from BitToPLC.Gestore import Gestore
from BitToPLC.HoverButton import HoverButton

x = 3
filename = 'Empty'
filename2 = 'Empty'
sheetIOlist = 'PLC IO LIST'
column_data = []
signal_types_list = []
signal_accepted_types = []
directory = 'M:/'
versione = '2.2'

def loadTypes():
    global signal_accepted_types
    with open('BitToPlc/txt/types.txt') as file:
        signal_accepted_types = file.readlines()
        print(signal_accepted_types)

# ----------------------------------------------------------------------------------------------------------------------------------------------------------------------

def openBitToPLC():

    def browseFiles():
        global filename
        global directory
        text.insert(INSERT, 'Wait...\n')
        text.see("end")
        filename = filedialog.askopenfilename(initialdir = directory, 
                                            title = "Seleziona una IOlist", 
                                            filetypes = (("Excel Files", "*.xlsx*"), ("Excel Files", "*.xlsm*"), ("all files", "*.*")))
        label_file_explorer.configure(text = "File aperto: " + filename)
        fileIoList = load_workbook(filename)
        variable.set('')
        opt['menu'].delete(0, 'end')
        directory = os.path.dirname(filename)

        file_log = open("log.txt", "a")
        file_log.write(time.ctime() + ' | ' + os.getlogin() + ': ' + filename + '\n')
        file_log.close()

        # nuova lista tendina
        nuovoMenu = []
        nuovoMenu = fileIoList.sheetnames
        for choice in nuovoMenu:
            opt['menu'].add_command(label=choice, command = tk._setit(variable, choice))
        text.insert(INSERT, 'Done.\n')
        text.see("end")

    def browseFiles2():
        global filename2
        global directory
        filename2 = filedialog.askopenfilename(initialdir = directory, 
                                            title = "Seleziona file excel BIT PLC", 
                                            filetypes = (("Excel Files", "*.xlsx*"), ("Excel Files", "*.xlsm*"), ("all files", "*.*")))
        label_file_explorer2.configure(text = "File aperto: " + filename2)
        directory = os.path.dirname(filename2)

    def inserisciDati():
        global sheetIOlist
        sheetIOlist = entry1.get()
        label_file_explorer4.configure(text = 'Figlio Excel: ' + sheetIOlist + '\n' + 'Lunghezza Massima Descrizioni: ' + spin1.get())

    def readIOlist():
        global column_data, ioList, signal_types_list, signal_accepted_types, x
        try:
            fileIoList = load_workbook(filename)
            ioList = fileIoList[sheetIOlist]
        except:
            text2.insert(INSERT, "Errore: (IOLIST) nessun file o sheet excel selezionati.\n")
            text2.see("end")
            file_log = open("log.txt", "a")
            file_log.write(time.ctime() + ' | ' + os.getlogin() + ': ' + 'Errore: (IOLIST) nessun file o sheet excel selezionati.' + '\n')
            file_log.close()        
        column_data = Gestore.getColumnData(ioList)
        descr_length = spin1.get()
        signal_types_list = SignalTypesArray()
        Gestore.readSignals(ioList, text2, x, column_data, descr_length, signal_types_list, signal_accepted_types)
        Gestore.printOp(text, signal_types_list)
        label_file_explorer5.configure(text = '\n')
        for i in signal_types_list:
            label_file_explorer5.configure(text = SignalTypeName(i) + ' =  ' + str(len(i)) + '\t\t')            

    def ElaboraDati():
        global filename, fileBit, bits, signal_types_list, filename2, column_data, ioList
        try:        
            fileBit = load_workbook(filename2)
            bits = fileBit["BIT"]
            bits = fileBit.active
        except:
            text2.insert(INSERT, "Errore: (FILE_BIT) nessun file selezionato.\n")
            text2.see("end")
            file_log = open("log.txt", "a")
            file_log.write(time.ctime() + ' | ' + os.getlogin() + ': ' + 'Errore: (IOLIST) nessun file o sheet excel selezionati.' + '\n')
            file_log.close()
        readIOlist()
        Gestore.Elaboration(filename2, fileBit, bits, text, text2, label_file_explorer3, signal_types_list)
        label_file_explorer3.configure(text = 'Done.')

    #costruzione di arrays per ogni tipo di segnale
    def SignalTypesArray():
        AI, AO, DI, DO, SAI, SAO, SDI, SDO, RAI, RAO, RDI, RDO, RSAI, RSAO, RSDI, RSDO, RTD = ([] for i in range(17)) #inizializzazioni di più arrays
        signal_types_list = [AI, AO, DI, DO, SAI, SAO, SDI, SDO, RAI, RAO, RDI, RDO, RSAI, RSAO, RSDI, RSDO, RTD]
        return signal_types_list
    
    #da usare quando serve stampare il nome di un array come stringa
    def SignalTypeName(array):
        name = [ i for i, j in locals().items() if j == array][0]
        return name
    
    '''
    def OpenNew():  # DA USARE PER AGGIUNGERE PAGINE GRAFICHE
        # OPEN WINDOW
        window = Tk()

        # FINESTRA
        window.title('SEID File Explorer')
        window.geometry("700x320")
        window.config(background = "black")

        # TITOLO SCHERMATA
        label_file_explorer = Label(window, text = "Import BITs to PLC", width = 80, height = 4, fg = "black", background = "#4bc973")   
        label_file_explorer2 = Label(window, text = "Import BITs to PLC", width = 80, height = 4, fg = "black", background = "#4bc973")
        label_file_explorer3 = Label(window, text = "", width = 80, height = 4, fg = "black", background = "#4bc973")
        label_file_explorer4 = Label(window, text = "", width = 80, height = 4, fg = "black", background = "#4bc973")
        # BUTTON
        button_esplora = Button(window, text = "IOList File", command = browseFiles, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')  
        button_esplora2 = Button(window, text = "BitPLC File", command = browseFiles2, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')
        button_exit = Button(window, text = "Annulla", command = OpenNew, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')
        button_elabora = Button(window, text = "Elabora", command = ElaboraDati, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')

        # Struttura
        label_file_explorer.grid(column = 2, row = 1)
        label_file_explorer2.grid(column = 2, row = 2)
        label_file_explorer3.grid(column = 2, row = 3)
        label_file_explorer4.grid(column = 2, row = 4)
        button_esplora.grid(column = 1, row = 1)
        button_esplora2.grid(column = 1, row = 2)
        button_elabora.grid(column = 1, row = 3)
        button_exit.grid(column = 1, row = 4)

        # loop window
        window.mainloop()
        '''

    # ----------------------------------------------------------------------------------------------------------------------------------------------------------------------

    # OPEN WINDOW
    window = Tk()

    #IMAGES
    img_folder = 'pkm'


    window.iconbitmap('BitToPlc/img/seid.ico')
    logo = PhotoImage(file='BitToPlc/img/SEIDN.jpg')
    image_menu = PhotoImage(file = 'BitToPlc/img/'+ img_folder + '/1.png')
    image_menu2 = PhotoImage(file = 'BitToPlc/img/'+ img_folder + '/2.png')
    image_menu3 = PhotoImage(file = 'BitToPlc/img/'+ img_folder + '/3.png')
    image_menu4 = PhotoImage(file = 'BitToPlc/img/'+ img_folder + '/7.png')
    image_menu5 = PhotoImage(file = 'BitToPlc/img/'+ img_folder + '/10.png')

    # FONT
    font1 = font.Font(family='Helvetica', size=8, weight='bold')

    # FINESTRA
    window.title('Tool: BIT to SPAC PLC')
    window.geometry("950x812")
    window.config(background = "black")

    # TITOLO SCHERMATA
    label_file_explorer = Label(window, text = "Nessun File Aperto", width = 80, height = 4, fg = "white", background = "#333333")   
    label_file_explorer2 = Label(window, text = "Nessun File Aperto", width = 80, height = 4, fg = "white", background = "#333333")
    label_file_explorer3 = Label(window, text = "", width = 80, height = 4, fg = "white", background = "#333333")
    label_file_explorer4 = Label(window, text = "", width = 80, height = 4, fg = "white", background = "#333333")
    label_file_explorer5 = Label(window, text = "", width = 80, height = 4, fg = "white", background = "#333333", anchor = CENTER)
    label_output = Label(window, text = "   Output", width = 134, fg = "white", background = "#333333", anchor = W)
    labelLogo = Label(window, image=logo)

    # BUTTON
    button_esplora = HoverButton(window, text = "          IOList File", command = browseFiles, image = image_menu3, compound=LEFT, width=150, height=45, bd=0, font=font1, bg='#292626', fg='#ffffff', activebackground='#436b45', activeforeground='#ffffff')
    button_esplora2 = HoverButton(window, text = "          BitPLC File", command = browseFiles2, image = image_menu2, compound=LEFT, width=150, height=45, bd=0, font=font1, bg='#292626', fg='#ffffff', activebackground='#436b45', activeforeground='#ffffff')
    button_exit = HoverButton(window, text = "Chiudi", command = exit, width=21, height=3, bd=0, font=font1, bg='#292626', fg='#ffffff', activebackground='#704242', activeforeground='#ffffff')
    button_elabora = HoverButton(window, text = "          Elabora", command = ElaboraDati, image = image_menu4, compound=LEFT, width=150, height=45, bd=0, font=font1, bg='#292626', fg='#ffffff', activebackground='#436b45', activeforeground='#ffffff')
    button_leggi = HoverButton(window, text = "          Lettura", command = readIOlist, image = image_menu, compound=LEFT, width=150, height=45, bd=0, font=font1, bg='#292626', fg='#ffffff', activebackground='#436b45', activeforeground='#ffffff')
    button_inserisci = HoverButton(window, text = "          Inserisci", command = inserisciDati, image = image_menu5, compound=LEFT, width=150, height=45, bd=0, font=font1, bg='#292626', fg='#ffffff', activebackground='#436b45', activeforeground='#ffffff')

    # Struttura
    label_file_explorer.grid(column = 2, row = 1)
    label_file_explorer2.grid(column = 2, row = 2)
    label_file_explorer3.grid(column = 2, row = 4)
    label_file_explorer4.grid(column = 2, row = 5)
    label_file_explorer5.grid(column = 2, row = 3)
    label_output.grid(column = 1, row = 7, columnspan=3)
    labelLogo.grid(column = 3, row = 5)
    button_esplora.grid(column = 1, row = 1)
    button_esplora2.grid(column = 1, row = 2)
    button_elabora.grid(column = 1, row = 4)
    button_exit.grid(column = 1, row = 5)
    button_leggi.grid(column = 1, row = 3)
    button_inserisci.grid(column = 3, row = 4)

    # spinbox
    spin_var = StringVar()
    spin1 = Spinbox(window, from_=0, to=50, width=18, bd=0, textvariable=spin_var, font=('Arial', 15), bg='#444444', fg='white', activebackground='#444444')
    spin1.grid(row=2, column=3)
    spin_var.set(18)

    # scrollText
    frame = Frame(window, width=385, height=300, bd=0, relief=SUNKEN)
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)
    xscrollbar = Scrollbar(frame, orient=HORIZONTAL)
    # xscrollbar.grid(row=1, column=0, sticky=E+W)
    yscrollbar = Scrollbar(frame)
    # yscrollbar.grid(row=0, column=1, sticky=N+S)
    text = Text(frame, width=115, height=24, wrap=NONE, bd=0, xscrollcommand=xscrollbar.set, yscrollcommand=yscrollbar.set, fg = "#4bc973", background = "black")
    text.grid(row=0, column=0, sticky= N + S + E +W)
    xscrollbar.config(command=text.xview)
    yscrollbar.config(command=text.yview)
    frame.grid(row=8, column=1, columnspan=3)
    text.insert(INSERT, 'ImportBitToPLC - Version:' + versione + '\n')

    frame2 = Frame(window, width=385, height=300, bd=0, relief=SUNKEN)
    frame2.grid_rowconfigure(0, weight=1)
    frame2.grid_columnconfigure(0, weight=1)
    xscrollbar2 = Scrollbar(frame2, orient=HORIZONTAL)
    # xscrollbar.grid(row=1, column=0, sticky=E+W)
    yscrollbar2 = Scrollbar(frame2)
    # yscrollbar.grid(row=0, column=1, sticky=N+S)
    text2 = Text(frame2, width=115, height=5, wrap=NONE, bd=0, xscrollcommand=xscrollbar.set, yscrollcommand=yscrollbar.set, fg = "red", background = "black")
    text2.grid(row=0, column=0, sticky= N + S + E +W)
    xscrollbar2.config(command=text2.xview)
    yscrollbar2.config(command=text2.yview)
    frame2.grid(row=6, column=1, columnspan=3)

    # ENTRY text
    name = StringVar()
    label11 = Label(window, text = "IOList Nome Foglio +\nLunghezza Massima Descrizioni", width=33, height=4, background = '#bfbfbf')
    label11.grid(column = 3, row = 1)
    entry1 = Entry(window, width = 25, font=('Arial', 13), textvariable = StringVar(), bd=0, background = '#bfbfbf')
    entry1.grid(column = 3, row = 3)
    entry1.insert(INSERT, 'PLC IO LIST')
    inserisciDati()

    # Menu a tendina
    OptionList = ['PLC IO LIST']
    variable = StringVar(window)
    variable.set(OptionList[0])
    opt = OptionMenu(window, variable, *OptionList)
    opt.config(font=('Helvetica', 8), width=20, height=1, bd=0, bg='#292626', fg='white', activebackground='#292626', activeforeground='#ffffff')
    opt["menu"].config(bd=0, bg='#292626', fg='white', activebackground='#595656', activeforeground='#ffffff')
    opt.focus()
    opt["menu"].focus()
    opt.grid(column = 3, row = 2)

    def callback(*args):
        # labelTest.configure(text="The selected item is {}".format(variable.get()))
        entry1.delete(0, 'end')
        entry1.insert(INSERT, str(variable.get()))
        inserisciDati()
        
    variable.trace("w", callback)
    # loop window
    loadTypes()
    window.mainloop()