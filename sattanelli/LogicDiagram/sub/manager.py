
from openpyxl import load_workbook
from sub.analog import AiBlock
from sub.analog import AnalogInput
from sub.digital_input import DigitalInput
from sub.digital_input import DiBlock


internal_link = '' #LogicDiagram/
row_first = 3
column_data = []
IO_list = None
logic_file = None
paramenters_sheet = None
parameters = None
line = 2



def getCell(x, y):
    return str(IO_list.cell(x, y).value)

def getCell2(x, y):
    return str(paramenters_sheet.cell(x, y).value)

def writeCell(x, y, value):
    logic_file.cell(x, y).value = value

class Manager:

    @staticmethod
    def getInternalLink():
        return internal_link

    @staticmethod
    def loadExcel(path, sheet_name):
        my_file = load_workbook(path, data_only=True)
        sheet = my_file[sheet_name]
        return sheet

    @staticmethod
    def combineTag(tag1, tag2, tag3, tag4):
        arr = [tag1, tag2, tag3, tag4]
        tag = ''
        for i in range(3):
            if arr[i] != 'None':
                tag = tag + arr[i]
        return tag

    @staticmethod
    def getColumnData(ioList):
        global column_data
        column_data = [0 for i in range(20)]
        y = 1
        while y != 50:
            if (str(ioList.cell(1, y).value) == '#Card'):
                column_data[0] = y
            if (str(ioList.cell(1, y).value) == '#TAG'):
                column_data[1] = y
            if (str(ioList.cell(1, y).value) == '#Descr'):
                column_data[2] = y
            if (str(ioList.cell(1, y).value) == '#Device'):
                column_data[3] = y
            if (str(ioList.cell(1, y).value) == '#LL'):
                column_data[4] = y
            if (str(ioList.cell(1, y).value) == '#L'):
                column_data[5] = y
            if (str(ioList.cell(1, y).value) == '#H'):
                column_data[6] = y
            if (str(ioList.cell(1, y).value) == '#HH'):
                column_data[7] = y
            if (str(ioList.cell(1, y).value) == '#Circuit'):
                column_data[8] = y
            if (str(ioList.cell(1, y).value) == '#DigAlarm'):
                column_data[9] = y
            if (str(ioList.cell(1, y).value) == '#Location'):
                column_data[10] = y
            if (str(ioList.cell(1, y).value) == '#Tag1'):
                column_data[11] = y
            if (str(ioList.cell(1, y).value) == '#Tag2'):
                column_data[12] = y
            if (str(ioList.cell(1, y).value) == '#Tag3'):
                column_data[13] = y
            if (str(ioList.cell(1, y).value) == '#Tag4'):
                column_data[14] = y
            if (str(ioList.cell(1, y).value) == '#Tag5'):
                column_data[15] = y
            if (str(ioList.cell(1, y).value) == '#Tag6'):
                column_data[16] = y
            if (str(ioList.cell(1, y).value) == '#Tag7'):
                column_data[17] = y
            if (str(ioList.cell(1, y).value) == '#Tag8'):
                column_data[18] = y
            y += 1
        return column_data

    @staticmethod
    def readIoList_AI(io_list):
        global IO_list
        global column_data
        IO_list = io_list
        signals = []; signal_type = ''
        i = 0
        Manager.getColumnData(io_list)
        while getCell(row_first+i, column_data[0]) != 'None' or getCell(row_first+i, column_data[2]) != 'None' or getCell(row_first+i, column_data[1]) != 'None':
            if getCell(row_first+i, column_data[0]) != 'None':
                signal_type = getCell(row_first+i, column_data[0])
            if 'AI' in signal_type and getCell(row_first+i, column_data[2]).lower() != 'spare':
                signals.append(AnalogInput( 'AI',  
                        getCell(row_first+i, column_data[1]),
                        getCell(row_first+i, column_data[2]),
                        getCell(row_first+i, column_data[3]),
                        getCell(row_first+i, column_data[7]),
                        getCell(row_first+i, column_data[6]),
                        getCell(row_first+i, column_data[5]),
                        getCell(row_first+i, column_data[4]),
                        getCell(row_first+i, column_data[8]),
                        getCell(row_first+i, column_data[10]),
                        Manager.combineTag(getCell(row_first+i, column_data[11]), getCell(row_first+i, column_data[12]),
                                           getCell(row_first+i, column_data[13]), getCell(row_first+i, column_data[14])),
                        Manager.combineTag(getCell(row_first+i, column_data[15]), getCell(row_first+i, column_data[16]),
                                           getCell(row_first+i, column_data[17]), getCell(row_first+i, column_data[18]))
                      ))  #creazione analogici
            i += 1
        return signals
    
    @staticmethod
    def readIoList_DI(io_list):
        global IO_list
        global column_data
        IO_list = io_list
        signals = []; signal_type = ''
        i = 0
        Manager.getColumnData(io_list)
        while getCell(row_first+i, column_data[0]) != 'None' or getCell(row_first+i, column_data[2]) != 'None' or getCell(row_first+i, column_data[1]) != 'None':
            if getCell(row_first+i, column_data[0]) != 'None':
                signal_type = getCell(row_first+i, column_data[0])
            if 'DI' in signal_type and getCell(row_first+i, column_data[2]).lower() != 'spare':
                signals.append(DigitalInput(    'DI',  
                        getCell(row_first+i, column_data[1]),
                        getCell(row_first+i, column_data[2]),
                        getCell(row_first+i, column_data[3]),
                        getCell(row_first+i, column_data[9]),
                        getCell(row_first+i, column_data[8]),
                        getCell(row_first+i, column_data[10]),
                        getCell(row_first+i, column_data[7]),
                        getCell(row_first+i, column_data[6]),
                        getCell(row_first+i, column_data[5]),
                        getCell(row_first+i, column_data[4]),
                        Manager.combineTag(getCell(row_first+i, column_data[11]), getCell(row_first+i, column_data[12]),
                                           getCell(row_first+i, column_data[13]), getCell(row_first+i, column_data[14])),
                        Manager.combineTag(getCell(row_first+i, column_data[15]), getCell(row_first+i, column_data[16]),
                                           getCell(row_first+i, column_data[17]), getCell(row_first+i, column_data[18]))
                    ))  #creazione digitali
            i += 1
        return signals
    
    @staticmethod
    def printArr(signal_list):
        i = 0
        while i < len(signal_list):
            signal_list[i].print()
            i += 1
    
    @staticmethod
    def printAllBlocks(signal_list):
        global parameters        
        parameters = Manager.readParameters()
        i = 0
        print('LEN' + str(len(signal_list)))
        while i < len(signal_list):
            data = AiBlock(signal_list[i], parameters, True, True, True, True, True, True, True, True, True, True, True)
            data.buildStructure()
            data.printBlock()
            print('\n@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@\n')
            i += 1
    
    @staticmethod
    def readParameters():
        global paramenters_sheet
        my_file = load_workbook(internal_link + 'BaseImportLogiciV3.xlsx')
        paramenters_sheet = my_file['Parametri']
        tipologia = []; codice = []; direzione = []; parameters = []; x = 3; y = 1
        while getCell2(x, y) != 'None':
            tipologia.append(getCell2(x, y+1))
            codice.append(getCell2(x, y))
            direzione.append(getCell2(x, y+2))
            x += 1
        parameters.append(codice)
        parameters.append(tipologia)
        parameters.append(direzione)
        return parameters
    
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def processDataBlock_AI(signal_list):
        global parameters        
        parameters = Manager.readParameters()
        total_blocks = []
        i = 0
        while i < len(signal_list):
            data = AiBlock(signal_list[i], parameters, False, True, False, True, True, True, False, False, False, False, False)
            data.processAnalogInputBlock()
            total_blocks.append(data.getFinalStructure())
            i += 1
        return total_blocks

    @staticmethod
    def processDataBlock_DI(signal_list):
        global parameters        
        parameters = Manager.readParameters()
        total_blocks = []
        i = 0
        while i < len(signal_list):
            data = DiBlock(signal_list[i], parameters)
            data.processDigitalInputBlock()
            total_blocks.append(data.getFinalStructure())
            i += 1
        return total_blocks

    @staticmethod
    def toExcel(total_blocks, sheet):
        global line, logic_file
        my_file = load_workbook(internal_link + 'BaseImportLogiciV3.xlsx')
        logic_file = my_file[sheet]
        k = 0; i = 2; x = 2; line = 2
        print('process')
        while k < len(total_blocks):
            print('█', end='')
            Manager.writeBlockToExcel(total_blocks[k])
            k += 1        
        #SAVE FILE
        my_file.save(internal_link + 'BaseImportLogiciV3.xlsx')
    
    @staticmethod
    def writeBlockToExcel(block_array):
        global line        
        i = 0; j = 0; x = line
        while i < (len(block_array)):
            j = 0
            while j < 32:
                row = block_array[i]
                writeCell(x, j+1, row[j])
                j += 1
            x += 1
            i += 1
        line = x

    @staticmethod
    def writeCommonToExcel(common_list, sheet):
        global logic_file
        my_file = load_workbook(internal_link + 'BaseImportLogiciV3.xlsx')
        logic_file = my_file[sheet]
        i = 0; j = 0
        while i < len(common_list):
            j = 0
            while j < 32:
                writeCell(i+2, j+1, common_list[i][j])
                j += 1
            i += 1
        #SAVE FILE
        my_file.save(internal_link + 'BaseImportLogiciV3.xlsx')