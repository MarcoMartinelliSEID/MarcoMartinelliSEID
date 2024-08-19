from curses.ascii import isdigit
from openpyxl import load_workbook
from item import Item
from typing import Type
import re

excel_1 = r'C:\Users\marcom\Documents\GitHub\sattanelli\BOM\Materials-List.xlsx'
excel_2 = r'C:\Users\marcom\Documents\GitHub\sattanelli\BOM\BOM.xlsx'

#excel file loading
def read_file(filename, sheetname):
    try:
        file = load_workbook(filename)
        sheet = file[sheetname]
    except:
        print('errore')
    return sheet, file

#creation of array from material-list
def fill_mat_array(excel):
    arr = []
    y = 2
    check = True #check the end of excel rows
    while check:
        item = Item(str(excel.cell(y,1).value), str(excel.cell(y,2).value), str(excel.cell(y,3).value),
                    str(excel.cell(y,4).value), str(excel.cell(y,5).value), str(excel.cell(y,6).value))
        if item.quantity == 'None' and item.item_name == 'None':
            check = False
        else:
            arr.append(item)
            y += 1
    return arr

def print_excel(arr, file, excel):
    y = 1; i = 0
    while i < len(arr):
        if arr[i].mark:
            excel.cell(y,1).value = str(arr[i].quantity)
            excel.cell(y,2).value = str(arr[i].item_name)
            excel.cell(y,3).value = str(arr[i].description_ita)
            excel.cell(y,4).value = str(arr[i].description_en)
            excel.cell(y,5).value = str(arr[i].code)
            excel.cell(y,6).value = str(arr[i].brand)
            y += 1 
        i += 1    
    file.save('BOM.xlsx') 

#count duplicates, remove and count
def count_duplicates(arr: list):
    count = 0; quantity = 0; i = 0
    for item in arr:
        while i < len(arr):
            if compare_item(arr[i], item):
                count += 1
                quantity = quantity + int(arr[i].quantity)
                '''
                if arr[i].code == 'OPT5096':
                    print(arr[i].item_name)
                    print('QTY: ' + str(arr[i].quantity) + ' contati: ' + str(quantity))
                '''
                if count > 1:
                    arr[i].mark = False
            i += 1  
        item.quantity = quantity
        item.item_name = re.sub(r'\d', 'x', item.item_name)
        i = 0; count = 0; quantity = 0
    return arr
    
def compare_item(item_1: Type[Item], item_2: Type[Item]):
    if item_1.product == item_2.product:
        return True
    else:
        return False

def print_console(arr):
    for pos in arr:
        if pos.mark:
            pos.print()

#-------------------------------------------------------------------------------------------------------

def main():
    mat_sheet, mat_file = read_file(excel_1, 'Distinta')
    bom_sheet, bom_file = read_file(excel_2, 'BOM2')
    arr = fill_mat_array(mat_sheet)
    print('\n\n-------------------------------------------------------------------\n\n')
    arr = count_duplicates(arr)
    print_console(arr)
    print_excel(arr, bom_file, bom_sheet)
    print('Saved')
    


#-------------------------------------------------------------------------------------------------------

if __name__ == "__main__":
    main()