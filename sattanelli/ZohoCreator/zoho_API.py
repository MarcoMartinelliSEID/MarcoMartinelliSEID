from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog
import time
import tkinter as tkinter
import tkinter.font as font
import requests
import json
import tablib
import pandas
import datetime
import re


prefix_id_csv = '334388'
prefix_id_excel = '3343883'
internal_link = '' #ZohoCreator/

def getInternalLink():
    return internal_link

def getPrefix(output):
    if output == 'csv':
        return prefix_id_csv
    elif output == 'xlsx':
        return prefix_id_excel

def cleanhtml(html):
    html = html.replace('<br />', '  |  ')
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', html)
    print(cleantext)
    return cleantext

def cleanJson(json):
    json = json.split("data\":",1)[1]
    json = json[0:-1]
    json = json.replace(prefix_id_excel, '@@@')
    #json.replace('{"display_value":', '')
    #json.replace('')
    return json


#°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°

def getRecord(text, url, token):
    headers = {
        'Authorization': 'Zoho-oauthtoken ' + token,
    }
    params = {}

    response = requests.get(url, headers=headers, data=params)
    text.insert(INSERT, response.text)
    json = cleanJson(response.text)
    pandas.read_json(json).to_excel(internal_link + 'files/' + 'serials.xlsx')
    fileIN = load_workbook(internal_link + 'files/' + 'serials.xlsx')
    sheet = fileIN["Sheet1"]
    sheet = fileIN.active
    sheet.delete_cols(1)
    fileIN.save(internal_link + 'files/' + 'serials.xlsx')

    json2 = str(pandas.read_excel(internal_link + 'files/' + 'serials.xlsx').to_json(orient='records'))
    json2 = json2.replace('@@@', prefix_id_excel)
    json2 = json2.replace('null', '""')
    json2 = json2.replace("': '", '":"')
    json2 = json2.replace("', '", '","')
    json2 = json2.replace("\"{'", '{"')
    json2 = json2.replace("'}\"", '"}')
    #print(json2)

def getRecordTest(text, url, token):
    headers = {
        'Authorization': 'Zoho-oauthtoken ' + token,
    }
    params = {}

    response = requests.get(url, headers=headers, data=params)
    print(response.text + 'FIGAAAA')
    json = cleanJson(response.text)
    pandas.read_json(json).to_excel(internal_link + 'files/' + 'serials.xlsx')
    fileIN = load_workbook(internal_link + 'files/' + 'serials.xlsx')
    sheet = fileIN["Sheet1"]
    sheet = fileIN.active
    sheet.delete_cols(1)
    fileIN.save(internal_link + 'files/' + 'serials.xlsx')   

    json2 = str(pandas.read_excel(internal_link + 'files/' + 'serials.xlsx').to_json(orient='records'))
    json2 = json2.replace('@@@', prefix_id_excel)
    json2 = json2.replace('null', '""')
    json2 = json2.replace("': '", '":"')
    json2 = json2.replace("', '", '","')
    json2 = json2.replace("\"{'", '{"')
    json2 = json2.replace("'}\"", '"}')
    #print(json2)    
    return response.text

def updateRecordTest(text, url, token, new_data):
    headers = {
    'Authorization': 'Zoho-oauthtoken ' + token,
    }
    new_data = new_data.replace('},', ',')
    new_data = new_data.replace('"ProductType":{', '')
    new_data = new_data.replace('"Commessa":{', '')
    print(new_data)
    data = new_data
    #data2 = '{"data":{"DateD_PO":"01-Mar-2021","ManufacturerID":"","Note":"","display_value":"Cabinet","ID":"3343883000000283017","display_value":"MI20010","ID":"3343883000000762011","DateD":"","ID":"3343883000000762018","Tag":"1JBAZi700","SerialN":"MI20010-01"}}'
    
    response = requests.patch(url, headers=headers, data=data)
    text.insert(INSERT, '\n' + response.text + '\n')


#°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°


def getRecordStatus(output, text, url, token):    

    headers = {
        'Authorization': 'Zoho-oauthtoken ' + token,
    }

    params = {}

    response = requests.get(url, headers=headers, data=params)

    #text.insert(INSERT, response.text + '\n')
    #print(response.text)

    if output == 'csv':
        jsonToCSV(response.text, prefix_id_csv)
    elif output == 'xlsx':
        jsonToXLSX(response.text, prefix_id_excel)
    

#°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°


def jsonToCSV(json, prefix):
    #json = '[{"Status":"FAT","ID":"33438830222"},{"Status":"Preparazione alla spedizione","ID":"3343883000567"},{"Status":"Ritirata","ID":"3343883000000322010"},{"Status":"Chiusa","ID":"3343883000000256064"},{"Status":"Spedita","ID":"3343883000000256058"},{"Status":"Commissioning","ID":"3343883000000256053"},{"Status":"Produzione","ID":"3343883000000256048"},{"Status":"Ingegneria","ID":"3343883000000256043"},{"Status":"Acquisita","ID":"3343883000000256038"},{"Status":"Bloccata","ID":"3343883000000256033"}]'
    json = json.split("data\":",1)[1]
    json = json[0:-1]
    json = json.replace(prefix, '')
    pandas.read_json(json).to_csv(internal_link + 'files/' + 'json.txt', index = False, sep = '@')

def jsonToXLSX(json, prefix):
    #json = '[{"Status":"FAT","ID":"33438830222"},{"Status":"Preparazione alla spedizione","ID":"3343883000567"},{"Status":"Ritirata","ID":"3343883000000322010"},{"Status":"Chiusa","ID":"3343883000000256064"},{"Status":"Spedita","ID":"3343883000000256058"},{"Status":"Commissioning","ID":"3343883000000256053"},{"Status":"Produzione","ID":"3343883000000256048"},{"Status":"Ingegneria","ID":"3343883000000256043"},{"Status":"Acquisita","ID":"3343883000000256038"},{"Status":"Bloccata","ID":"3343883000000256033"}]'
    json = json.split("data\":",1)[1]
    json = json[0:-1]
    json = json.replace(prefix, '')
    pandas.read_json(json).to_excel(internal_link + 'files/' + 'json.xlsx')


#°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°


def getSimpleRecord(text, url, token):    

    headers = {
        'Authorization': 'Zoho-oauthtoken ' + token,
    }
    params = {}

    response = requests.get(url, headers=headers, data=params)
    json = response.text
    #text.insert(INSERT, '%\nGET RECORD:\n' + json + '\n%\n')
    print(json)

    #ricavo ID dal json
    try:
        id_record = json.split('"ID":"',1)[1]
    except IndexError:
        text.insert(INSERT, '\nERROR: ID Job not found\n')
    json_splitted = id_record.split('"')
    id_record = json_splitted[0]
    print(id_record)
    #text.insert(INSERT, id_record + '\n')

    #ricavo le Note dal json
    note = json.split('"Note":"',1)[1]
    json_splitted = note.split('"')
    note = json_splitted[0]
    print(note)
    #text.insert(INSERT, note + '\n')    

    return id_record, note

def getCompleteRecord(text, labels, url, token):    

    headers = {
        'Authorization': 'Zoho-oauthtoken ' + token,
    }
    params = {}

    response = requests.get(url, headers=headers, data=params)
    json = response.text
    text.insert(INSERT, '%\nGET RECORD:\n' + json + '\n%\n')
    labels[0].configure(text = '     ' + str(getJsonData(json, 'JobID')))
    labels[1].configure(text = '     ' + str(getJsonData(json, 'PrjStatus":{"display_value')))
    labels[2].configure(text = '     ' + str(getJsonData(json, 'PM":{"display_value')))
    labels[3].configure(text = '     ' + str(getJsonData(json, 'PE":{"display_value')))    
    labels[4].configure(text = '     ' + str(cleanhtml(str(getJsonData(json, 'Note')))))

def getJsonData(json, value):
    try:
        data = (((json.split('"' + value + '":"',1)[1]).split('"'))[0])
    except IndexError:
        data = ''
    return data



#°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°


def updateRecord(text, url, new_status, status_name, note_record, token):
    headers = {
    'Authorization': 'Zoho-oauthtoken ' + token,
    }
    new_note = '<div>' + str(getDate()) + ' ' + status_name + '<br /><\/div>'
    new_note = note_record + new_note

    data = '{"data":{"CustPE":"","PE":"","PrjStatus":"' + new_status + '","Note":"' + new_note + '"}' + '}'
    #data2 = '{"data":{"CustPE":"","PE":"","PrjStatus":"3343883000000473066","Note":"<div>26-ttAug-2020 FAT<br /><\/div>","ID":"3343883000000675003","CustPM":"","PM":"3343883000000228655","JobID":"IU20030"}' + '}'
    
    response = requests.patch(url, headers=headers, data=data)
    result = str(response.text).split('"message":"',1)[1]
    text.insert(INSERT, '\nUPDATE RECORD: ' + result[0:-2] + '\n')

def getDate():
    date = datetime.datetime.now()
    format_date = (date.strftime("%"+"d") + '-' + date.strftime("%"+"b") + '-' + date.strftime("%Y"))
    return format_date 

#°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°
    
