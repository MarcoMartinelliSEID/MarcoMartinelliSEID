#----------------------------------------------------------------------------------------------------------------------------------------------------------------------
#SEID Srl
#Progetto: ModbusDB Generator
#Versione: 1.0
#Data: 10/06/2020
#Aggiornamenti: xxx
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------

from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog
import time
from BitToPLC import Segnale
from BitToPLC import Gestore
import tkinter as tkinter
import os

filename = 'Empty'

def openModbusDB():

    def runModbusDB():
        print("Processing...")
        label_elabora.configure(text = "Elaborazione in corso...")

        #lettura file.xlsx
        global filename
        fileIN = load_workbook(filename)
        sheet = fileIN["Foglio1"]

        #creazione file.awl
        fileOUT = open("SiemensSource.awl", "w+") #C:\\ ... tra le virgolette, per accedere a file in altri folder

        #+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        fileOUT.write("DATA_BLOCK db" + str(sheet.cell(2,2).value) + "\n")
        fileOUT.write("TITLE =" + "\n")
        fileOUT.write("VERSION : 0.1" + "\n") # + input())


        fileOUT.write("\n\n  STRUCT" + "\n")
        fileOUT.write("   ANALOG : STRUCT" + "\n")
        x = 4
        y = 2
        while str.upper(sheet.cell(x, y-1).value) == "ANALOG":  #accetta sia maiuscolo che minuscolo
            fileOUT.write("    " + "r" + str(sheet.cell(x, y).value) + " : REAL ;" + "\n")
            x += 1
        fileOUT.write("   END_STRUCT ;" + "\n")



        fileOUT.write("   STATUS : STRUCT" + "\n")
        while str.upper(sheet.cell(x, y-1).value) == "STATUS":  #accetta sia maiuscolo che minuscolo
            fileOUT.write("    " + "r" + str(sheet.cell(x, y).value) + " : INT ;" + "\n")
            x += 1
        fileOUT.write("   END_STRUCT ;" + "\n")



        fileOUT.write("   DIGITAL : STRUCT" + "\n")
        while sheet.cell(x, y-1).value is not None and str.upper(sheet.cell(x, y-1).value) == "DIGITAL":  #accetta sia maiuscolo che minuscolo
            fileOUT.write("    " + "r" + str(sheet.cell(x, y).value) + ' : "UDT_DCS" ;' + "\n")
            x += 1
        fileOUT.write("   END_STRUCT ;" + "\n")
        fileOUT.write("  END_STRUCT ;" + "\n")

        fileOUT.write("\n  BEGIN" + "\n")
        fileOUT.write("  END_DATA_BLOCK" + "\n")

        print("Done.")
        label_elabora.configure(text = "Operazione Completata")
        label_output.configure(text = "File Output Pronto")

    #--------------------------------------------------------------------------------------------------------------------------------------------------------------------

    def browseFiles():
        global filename
        filename = filedialog.askopenfilename(initialdir = "/", 
                                            title = "Seleziona il file excel", 
                                            filetypes = (("Excel Files", "*.xlsx*"), ("Excel Files", "*.xlsm*"), ("all files", "*.*")))

        label_file_explorer.configure(text = "File aperto: " + filename)
        label_elabora.configure(text = "Pronto")

    def openOutputFile():
        os.system("start " + 'SiemensSource.awl')

    
    #OPEN WINDOW
    ModbusDB = Tk()
    ModbusDB.iconbitmap('img/seid.ico')
    logo = PhotoImage(file="img/SEIDN.jpg")
            
    #FINESTRA
    ModbusDB.title('SEID File Explorer')
    ModbusDB.geometry("700x300")
    ModbusDB.config(background = "black")

    # TITOLO SCHERMATA
    label_file_explorer = Label(ModbusDB, text = "Nessun File Aperto", width = 80, height = 4, fg = "white", background = "#404040")   
    label_elabora = Label(ModbusDB, text = "Attesa", width = 80, height = 4, fg = "white", background = "#404040")
    label_output = Label(ModbusDB, text = "", width = 80, height = 4, fg = "white", background = "#404040")
    label_exit = Label(ModbusDB, text = "Exit", width = 80, height = 4, fg = "white", background = "#404040")
    #BUTTON
    button_esplora = Button(ModbusDB, text = "Open File", command = browseFiles, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')  
    button_elabora = Button(ModbusDB, text = "Elabora", command = runModbusDB, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')
    button_output = Button(ModbusDB, text = "Open Output File", command = openOutputFile, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')
    button_exit = Button(ModbusDB, text = "Chiudi", command = exit, width=20, height=2, bg='#292626', fg='#ffffff', activebackground='#292626', activeforeground='#ffffff')

    #Struttura
    label_file_explorer.grid(column = 2, row = 1)
    label_elabora.grid(column = 2, row = 2)
    label_output.grid(column = 2, row = 3)
    label_exit.grid(column = 2, row = 4)
    button_esplora.grid(column = 1, row = 1)
    button_elabora.grid(column = 1, row = 2)
    button_output.grid(column = 1, row = 3)
    button_exit.grid(column = 1, row = 4)

    #loop window
    ModbusDB.mainloop()

            
        